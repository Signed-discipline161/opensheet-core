"""Benchmark: opensheet_core vs openpyxl for reading XLSX files."""

import os
import sys
import time
import tempfile
import tracemalloc

import openpyxl
import opensheet_core


def generate_test_file(path, rows, cols):
    """Generate a test XLSX file using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark"

    # Header
    ws.append([f"col_{i}" for i in range(cols)])

    # Data rows with mixed types
    for r in range(rows):
        row = []
        for c in range(cols):
            match c % 4:
                case 0:
                    row.append(f"text_{r}_{c}")
                case 1:
                    row.append(r * cols + c)
                case 2:
                    row.append((r * cols + c) * 0.123)
                case 3:
                    row.append(r % 2 == 0)
        ws.append(row)

    wb.save(path)
    return os.path.getsize(path)


def bench_openpyxl_read(path):
    """Read all cells with openpyxl."""
    tracemalloc.start()
    t0 = time.perf_counter()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        _ = list(row)
    wb.close()

    elapsed = time.perf_counter() - t0
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    return elapsed, peak, row_count


def bench_opensheet_read(path):
    """Read all cells with opensheet_core."""
    tracemalloc.start()
    t0 = time.perf_counter()

    rows = opensheet_core.read_sheet(path)
    row_count = len(rows)

    elapsed = time.perf_counter() - t0
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    return elapsed, peak, row_count


def format_mem(bytes):
    if bytes < 1024:
        return f"{bytes} B"
    elif bytes < 1024 * 1024:
        return f"{bytes / 1024:.1f} KB"
    else:
        return f"{bytes / (1024 * 1024):.1f} MB"


def run_benchmark(rows, cols, runs=3):
    print(f"\n{'='*60}")
    print(f"Benchmark: {rows:,} rows x {cols} cols ({rows * cols:,} cells)")
    print(f"{'='*60}")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    try:
        file_size = generate_test_file(path, rows, cols)
        print(f"File size: {format_mem(file_size)}")
        print()

        # Warm up
        bench_opensheet_read(path)
        bench_openpyxl_read(path)

        # Benchmark opensheet_core
        os_times, os_mems = [], []
        for _ in range(runs):
            t, m, _ = bench_opensheet_read(path)
            os_times.append(t)
            os_mems.append(m)

        # Benchmark openpyxl
        op_times, op_mems = [], []
        for _ in range(runs):
            t, m, _ = bench_openpyxl_read(path)
            op_times.append(t)
            op_mems.append(m)

        os_avg = sum(os_times) / len(os_times)
        op_avg = sum(op_times) / len(op_times)
        os_mem = sum(os_mems) / len(os_mems)
        op_mem = sum(op_mems) / len(op_mems)

        speedup = op_avg / os_avg if os_avg > 0 else float("inf")
        mem_ratio = op_mem / os_mem if os_mem > 0 else float("inf")

        print(f"  {'Library':<20} {'Time (avg)':<15} {'Peak Memory':<15}")
        print(f"  {'-'*50}")
        print(f"  {'opensheet_core':<20} {os_avg*1000:.1f} ms{'':<8} {format_mem(os_mem):<15}")
        print(f"  {'openpyxl':<20} {op_avg*1000:.1f} ms{'':<8} {format_mem(op_mem):<15}")
        print()
        print(f"  Speed:  opensheet_core is {speedup:.1f}x faster")
        print(f"  Memory: opensheet_core uses {mem_ratio:.1f}x less peak memory")

        return {
            "rows": rows,
            "cols": cols,
            "opensheet_time": os_avg,
            "openpyxl_time": op_avg,
            "opensheet_mem": os_mem,
            "openpyxl_mem": op_mem,
            "speedup": speedup,
            "mem_ratio": mem_ratio,
        }
    finally:
        os.unlink(path)


def main():
    print("OpenSheet Core vs openpyxl — Read Benchmark")
    print(f"opensheet_core {opensheet_core.__version__}")
    print(f"openpyxl {openpyxl.__version__}")
    print(f"Python {sys.version.split()[0]}")

    configs = [
        (1_000, 10),
        (10_000, 10),
        (50_000, 10),
        (100_000, 10),
        (10_000, 50),
    ]

    results = []
    for rows, cols in configs:
        result = run_benchmark(rows, cols)
        results.append(result)

    print(f"\n{'='*60}")
    print("Summary")
    print(f"{'='*60}")
    print(f"  {'Config':<20} {'Speedup':<12} {'Mem Savings':<12}")
    print(f"  {'-'*44}")
    for r in results:
        config = f"{r['rows']:,} x {r['cols']}"
        print(f"  {config:<20} {r['speedup']:.1f}x{'':<8} {r['mem_ratio']:.1f}x")


if __name__ == "__main__":
    main()
