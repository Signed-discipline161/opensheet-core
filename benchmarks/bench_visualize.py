#!/usr/bin/env python3
"""
OpenSheet Core Benchmark Visualization

Runs benchmarks across multiple dataset sizes, collects results, and generates
SVG charts comparing opensheet_core vs openpyxl performance.

Usage:
    # Run benchmarks and generate charts
    python benchmarks/bench_visualize.py

    # Generate charts from existing cached results (skip benchmarking)
    python benchmarks/bench_visualize.py --no-run

    # Custom runs per configuration
    python benchmarks/bench_visualize.py --runs 3

    # Quick mode (1 run per config, for testing the script itself)
    python benchmarks/bench_visualize.py --quick
"""

import argparse
import json
import os
import platform
import sys
import tempfile
from datetime import datetime, timezone

# Ensure the benchmarks directory is on sys.path so bench_utils can be imported
# regardless of the working directory.
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

from bench_utils import bench_pair, format_bytes, format_time, generate_row

try:
    import openpyxl
except ImportError:
    print("openpyxl is required for benchmarking: pip install openpyxl")
    sys.exit(1)

try:
    import matplotlib

    matplotlib.use("Agg")  # Non-interactive backend for headless chart generation
    import matplotlib.pyplot as plt
except ImportError:
    print("matplotlib is required for chart generation: pip install matplotlib")
    sys.exit(1)

import opensheet_core

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = os.path.dirname(_HERE)
ASSETS_DIR = os.path.join(PROJECT_ROOT, "docs", "assets")
RESULTS_PATH = os.path.join(ASSETS_DIR, "benchmark_results.json")

# ---------------------------------------------------------------------------
# Benchmark configurations: (rows, cols) pairs
# ---------------------------------------------------------------------------

CONFIGS = [
    (1_000, 10),
    (10_000, 10),
    (50_000, 10),
    (100_000, 10),
]

# ---------------------------------------------------------------------------
# Chart colors
# ---------------------------------------------------------------------------

COLOR_OPENSHEET = "#2563eb"  # blue
COLOR_OPENPYXL = "#dc2626"  # red

# ---------------------------------------------------------------------------
# Write / read functions (same as existing benchmark scripts)
# ---------------------------------------------------------------------------


def write_opensheet(path, rows, cols):
    with opensheet_core.XlsxWriter(path) as w:
        w.add_sheet("Benchmark")
        w.write_row([f"col_{i}" for i in range(cols)])
        for r in range(rows):
            w.write_row(generate_row(r, cols))


def write_openpyxl(path, rows, cols):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Benchmark")
    ws.append([f"col_{i}" for i in range(cols)])
    for r in range(rows):
        ws.append(generate_row(r, cols))
    wb.save(path)


def read_opensheet(path):
    rows = opensheet_core.read_sheet(path)
    _ = len(rows)


def read_openpyxl(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    _ = len(all_rows)
    wb.close()


# ---------------------------------------------------------------------------
# Benchmark runner
# ---------------------------------------------------------------------------


def _generate_test_file(path, rows, cols):
    """Generate a test XLSX file using openpyxl (used as the read target)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark"
    ws.append([f"col_{i}" for i in range(cols)])
    for r in range(rows):
        ws.append(generate_row(r, cols))
    wb.save(path)
    return os.path.getsize(path)


def run_all_benchmarks(configs, runs=5):
    """Run read and write benchmarks for every (rows, cols) configuration.

    Returns a dict with metadata and a list of per-configuration result dicts.
    """
    print(f"Running benchmarks: {len(configs)} configs, {runs} interleaved runs each")
    print(f"  opensheet_core  {opensheet_core.__version__}")
    print(f"  openpyxl        {openpyxl.__version__}")
    print(f"  Python          {sys.version.split()[0]}")
    print()

    all_results = []

    for idx, (rows, cols) in enumerate(configs, 1):
        label = f"[{idx}/{len(configs)}] {rows:,} rows x {cols} cols"
        print(f"{label} ...")

        # Temporary files for write benchmarks
        fd_os, os_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd_os)
        fd_op, op_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd_op)

        # Temporary file for read benchmarks (openpyxl-generated for fairness)
        fd_rd, read_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd_rd)

        try:
            # -- Write benchmark --
            # Warm up
            write_opensheet(os_path, min(rows, 100), cols)
            write_openpyxl(op_path, min(rows, 100), cols)

            os_wr, op_wr = bench_pair(
                write_opensheet, (os_path, rows, cols),
                write_openpyxl, (op_path, rows, cols),
                runs=runs,
            )

            # -- Read benchmark --
            _generate_test_file(read_path, rows, cols)
            # Warm up
            read_opensheet(read_path)
            read_openpyxl(read_path)

            os_rr, op_rr = bench_pair(
                read_opensheet, (read_path,),
                read_openpyxl, (read_path,),
                runs=runs,
            )

            result = {
                "rows": rows,
                "cols": cols,
                "read_opensheet_time": os_rr.min_time,
                "read_openpyxl_time": op_rr.min_time,
                "read_opensheet_mem": os_rr.median_mem,
                "read_openpyxl_mem": op_rr.median_mem,
                "write_opensheet_time": os_wr.min_time,
                "write_openpyxl_time": op_wr.min_time,
                "write_opensheet_mem": os_wr.median_mem,
                "write_openpyxl_mem": op_wr.median_mem,
            }
            all_results.append(result)

            read_speedup = op_rr.min_time / os_rr.min_time if os_rr.min_time > 0 else float("inf")
            write_speedup = op_wr.min_time / os_wr.min_time if os_wr.min_time > 0 else float("inf")
            print(f"  Read:  {format_time(os_rr.min_time)} vs {format_time(op_rr.min_time)}  ({read_speedup:.1f}x)")
            print(f"  Write: {format_time(os_wr.min_time)} vs {format_time(op_wr.min_time)}  ({write_speedup:.1f}x)")

        finally:
            for p in (os_path, op_path, read_path):
                if os.path.exists(p):
                    os.unlink(p)

    data = {
        "metadata": {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "opensheet_core_version": opensheet_core.__version__,
            "openpyxl_version": openpyxl.__version__,
            "python_version": sys.version.split()[0],
            "platform": platform.platform(),
            "machine": platform.machine(),
            "runs_per_config": runs,
        },
        "results": all_results,
    }
    return data


# ---------------------------------------------------------------------------
# Chart generation
# ---------------------------------------------------------------------------


def _row_labels(results):
    """Generate x-axis labels like '1K', '10K', '50K', '100K'."""
    labels = []
    for r in results:
        rows = r["rows"]
        if rows >= 1_000_000:
            labels.append(f"{rows // 1_000_000}M")
        elif rows >= 1_000:
            labels.append(f"{rows // 1_000}K")
        else:
            labels.append(str(rows))
    return labels


def _save_figure(fig, name):
    """Save a figure as SVG and PNG to the assets directory."""
    os.makedirs(ASSETS_DIR, exist_ok=True)
    svg_path = os.path.join(ASSETS_DIR, f"{name}.svg")
    png_path = os.path.join(ASSETS_DIR, f"{name}.png")
    fig.savefig(svg_path, format="svg", bbox_inches="tight")
    fig.savefig(png_path, format="png", bbox_inches="tight", dpi=150)
    print(f"  Saved {svg_path}")
    print(f"  Saved {png_path}")
    plt.close(fig)


def _apply_chart_style(ax, title, ylabel):
    """Apply consistent styling to a chart axis."""
    ax.set_title(title, fontsize=14, fontweight="bold", pad=12)
    ax.set_ylabel(ylabel, fontsize=11)
    ax.set_xlabel("Dataset size (rows x 10 cols)", fontsize=11)
    ax.legend(fontsize=10, loc="upper left")
    ax.grid(axis="y", alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)


def generate_read_time_chart(results):
    """Generate grouped bar chart for read times."""
    labels = _row_labels(results)
    os_times = [r["read_opensheet_time"] for r in results]
    op_times = [r["read_openpyxl_time"] for r in results]

    fig, ax = plt.subplots(figsize=(8, 5))
    x = range(len(labels))
    width = 0.35

    bars_os = ax.bar([i - width / 2 for i in x], os_times, width, label="opensheet_core", color=COLOR_OPENSHEET)
    bars_op = ax.bar([i + width / 2 for i in x], op_times, width, label="openpyxl", color=COLOR_OPENPYXL)

    # Add value labels on bars
    for bar in bars_os:
        height = bar.get_height()
        ax.annotate(
            format_time(height),
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 4),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=8,
        )
    for bar in bars_op:
        height = bar.get_height()
        ax.annotate(
            format_time(height),
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 4),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=8,
        )

    ax.set_xticks(list(x))
    ax.set_xticklabels(labels)
    _apply_chart_style(ax, "Read Performance: opensheet_core vs openpyxl", "Time (seconds)")

    _save_figure(fig, "bench_read_time")


def generate_write_time_chart(results):
    """Generate grouped bar chart for write times."""
    labels = _row_labels(results)
    os_times = [r["write_opensheet_time"] for r in results]
    op_times = [r["write_openpyxl_time"] for r in results]

    fig, ax = plt.subplots(figsize=(8, 5))
    x = range(len(labels))
    width = 0.35

    bars_os = ax.bar([i - width / 2 for i in x], os_times, width, label="opensheet_core", color=COLOR_OPENSHEET)
    bars_op = ax.bar([i + width / 2 for i in x], op_times, width, label="openpyxl", color=COLOR_OPENPYXL)

    for bar in bars_os:
        height = bar.get_height()
        ax.annotate(
            format_time(height),
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 4),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=8,
        )
    for bar in bars_op:
        height = bar.get_height()
        ax.annotate(
            format_time(height),
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 4),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=8,
        )

    ax.set_xticks(list(x))
    ax.set_xticklabels(labels)
    _apply_chart_style(ax, "Write Performance: opensheet_core vs openpyxl", "Time (seconds)")

    _save_figure(fig, "bench_write_time")


def generate_memory_chart(results):
    """Generate grouped bar chart for memory usage (read and write)."""
    labels = _row_labels(results)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

    x = range(len(labels))
    width = 0.35

    # Read memory
    os_read_mem = [r["read_opensheet_mem"] / (1024 * 1024) for r in results]
    op_read_mem = [r["read_openpyxl_mem"] / (1024 * 1024) for r in results]

    ax1.bar([i - width / 2 for i in x], os_read_mem, width, label="opensheet_core", color=COLOR_OPENSHEET)
    ax1.bar([i + width / 2 for i in x], op_read_mem, width, label="openpyxl", color=COLOR_OPENPYXL)
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(labels)
    _apply_chart_style(ax1, "Read Memory (RSS Delta)", "Memory (MB)")

    # Write memory
    os_write_mem = [r["write_opensheet_mem"] / (1024 * 1024) for r in results]
    op_write_mem = [r["write_openpyxl_mem"] / (1024 * 1024) for r in results]

    ax2.bar([i - width / 2 for i in x], os_write_mem, width, label="opensheet_core", color=COLOR_OPENSHEET)
    ax2.bar([i + width / 2 for i in x], op_write_mem, width, label="openpyxl", color=COLOR_OPENPYXL)
    ax2.set_xticks(list(x))
    ax2.set_xticklabels(labels)
    _apply_chart_style(ax2, "Write Memory (RSS Delta)", "Memory (MB)")

    fig.suptitle("Memory Usage: opensheet_core vs openpyxl", fontsize=15, fontweight="bold", y=1.02)
    fig.tight_layout()

    _save_figure(fig, "bench_memory")


def generate_speedup_chart(results):
    """Generate a line chart showing speedup ratio across dataset sizes."""
    labels = _row_labels(results)
    read_speedups = []
    write_speedups = []

    for r in results:
        rs = r["read_openpyxl_time"] / r["read_opensheet_time"] if r["read_opensheet_time"] > 0 else 0
        ws = r["write_openpyxl_time"] / r["write_opensheet_time"] if r["write_opensheet_time"] > 0 else 0
        read_speedups.append(rs)
        write_speedups.append(ws)

    fig, ax = plt.subplots(figsize=(8, 5))
    x = range(len(labels))

    ax.plot(list(x), read_speedups, "o-", color=COLOR_OPENSHEET, linewidth=2, markersize=8, label="Read speedup")
    ax.plot(list(x), write_speedups, "s--", color=COLOR_OPENPYXL, linewidth=2, markersize=8, label="Write speedup")

    # Add value labels
    for i, (rs, ws) in enumerate(zip(read_speedups, write_speedups)):
        ax.annotate(f"{rs:.1f}x", (i, rs), textcoords="offset points", xytext=(0, 10), ha="center", fontsize=9)
        ax.annotate(f"{ws:.1f}x", (i, ws), textcoords="offset points", xytext=(0, -14), ha="center", fontsize=9)

    ax.axhline(y=1.0, color="gray", linestyle=":", alpha=0.5)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels)
    ax.set_ylabel("Speedup (x faster than openpyxl)", fontsize=11)
    ax.set_xlabel("Dataset size (rows x 10 cols)", fontsize=11)
    ax.set_title("opensheet_core Speedup Over openpyxl", fontsize=14, fontweight="bold", pad=12)
    ax.legend(fontsize=10)
    ax.grid(axis="y", alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    _save_figure(fig, "bench_speedup")


def generate_all_charts(data):
    """Generate all charts from benchmark data."""
    results = data["results"]
    print("\nGenerating charts ...")
    generate_read_time_chart(results)
    generate_write_time_chart(results)
    generate_memory_chart(results)
    generate_speedup_chart(results)
    print("\nAll charts saved to:", ASSETS_DIR)


# ---------------------------------------------------------------------------
# JSON persistence
# ---------------------------------------------------------------------------


def save_results(data):
    """Save benchmark results as JSON."""
    os.makedirs(ASSETS_DIR, exist_ok=True)
    with open(RESULTS_PATH, "w") as f:
        json.dump(data, f, indent=2)
    print(f"\nResults saved to {RESULTS_PATH}")


def load_results():
    """Load cached benchmark results from JSON."""
    if not os.path.exists(RESULTS_PATH):
        print(f"No cached results found at {RESULTS_PATH}")
        print("Run without --no-run first to generate benchmark data.")
        sys.exit(1)
    with open(RESULTS_PATH) as f:
        data = json.load(f)
    meta = data.get("metadata", {})
    print(f"Loaded cached results from {RESULTS_PATH}")
    print(f"  Timestamp:         {meta.get('timestamp', 'unknown')}")
    print(f"  opensheet_core:    {meta.get('opensheet_core_version', 'unknown')}")
    print(f"  openpyxl:          {meta.get('openpyxl_version', 'unknown')}")
    print(f"  Python:            {meta.get('python_version', 'unknown')}")
    print(f"  Platform:          {meta.get('platform', 'unknown')}")
    print(f"  Runs per config:   {meta.get('runs_per_config', 'unknown')}")
    return data


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Run benchmarks and generate visualization charts for opensheet_core vs openpyxl."
    )
    parser.add_argument(
        "--no-run",
        action="store_true",
        help="Skip running benchmarks; generate charts from cached JSON results only.",
    )
    parser.add_argument(
        "--runs",
        type=int,
        default=5,
        help="Number of interleaved runs per configuration (default: 5).",
    )
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Quick mode: 1 run per configuration (for testing the script).",
    )
    args = parser.parse_args()

    if args.quick:
        args.runs = 1

    if args.no_run:
        data = load_results()
    else:
        data = run_all_benchmarks(CONFIGS, runs=args.runs)
        save_results(data)

    generate_all_charts(data)


if __name__ == "__main__":
    main()
